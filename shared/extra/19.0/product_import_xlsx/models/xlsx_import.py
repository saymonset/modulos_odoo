import base64
import io
import logging
import os
import zipfile
from collections import defaultdict
from datetime import datetime

from odoo import _, api, fields, models
from odoo.exceptions import UserError, ValidationError

_logger = logging.getLogger(__name__)

try:
    import openpyxl
except ImportError:
    _logger.warning("openpyxl no está instalado. Instálelo con: pip install openpyxl")


class ProductXlsxImport(models.Model):
    _name = 'product.xlsx.import'
    _description = 'Importación de productos desde XLSX'
    _rec_name = 'name'
    _order = 'create_date desc'

    name = fields.Char(
        string='Nombre',
        required=True,
        default=lambda self: _('Importación %s') % datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
    )
    xlsx_file = fields.Binary(
        string='Archivo XLSX',
        required=True,
        attachment=True,
    )
    xlsx_filename = fields.Char(string='Nombre del archivo XLSX')
    imagenes_zip = fields.Binary(
        string='ZIP de imágenes',
        attachment=True,
        help='Archivo ZIP con las imágenes referenciadas en la columna "Imagen (nombre archivo)"',
    )
    imagenes_zip_filename = fields.Char(string='Nombre del ZIP')
    categ_id = fields.Many2one(
        'product.category',
        string='Categoría de producto',
        help='Categoría que se asignará a todos los productos importados. Vacío = usa la raíz.',
    )
    state = fields.Selection([
        ('draft', 'Borrador'),
        ('importing', 'Importando'),
        ('done', 'Completado'),
        ('error', 'Error'),
    ], string='Estado', default='draft', readonly=True)
    log = fields.Text(string='Log de importación', readonly=True)
    created_products = fields.Integer(string='Productos creados', readonly=True)
    created_variants = fields.Integer(string='Variantes creadas', readonly=True)
    created_attributes = fields.Integer(string='Atributos creados', readonly=True)
    skipped_count = fields.Integer(string='Omitidos', readonly=True)
    error_count = fields.Integer(string='Errores', readonly=True)

    def action_import(self):
        if not self.xlsx_file:
            raise UserError(_('Debe seleccionar un archivo XLSX'))
        self.write({'state': 'importing', 'log': '', 'created_products': 0,
                     'created_variants': 0, 'created_attributes': 0,
                     'skipped_count': 0, 'error_count': 0})
        try:
            result = self._run_import()
            self.write({
                'state': 'done',
                'log': result.get('log', ''),
                'created_products': result.get('created_products', 0),
                'created_variants': result.get('created_variants', 0),
                'created_attributes': result.get('created_attributes', 0),
                'skipped_count': result.get('skipped_count', 0),
                'error_count': result.get('error_count', 0),
            })
        except Exception as e:
            _logger.exception('Error en importación XLSX')
            self.write({'state': 'error', 'log': (self.log or '') + f'\nERROR CRÍTICO: {e}'})
            raise UserError(_('Error durante la importación: %s') % e)

    def _run_import(self):
        log_lines = []
        stats = {'created_products': 0, 'created_variants': 0,
                 'created_attributes': 0, 'skipped_count': 0, 'error_count': 0}
        images = {}
        zip_content = self.imagenes_zip and base64.b64decode(self.imagenes_zip)
        if zip_content:
            images = self._extract_zip_images(zip_content)
        xlsx_content = base64.b64decode(self.xlsx_file)
        groups = self._parse_xlsx(xlsx_content)
        log_lines.append(f'📂 Archivo: {self.xlsx_filename or "XLSX"}')
        log_lines.append(f'📦 Productos a importar: {len(groups)}')
        if images:
            log_lines.append(f'🖼️  Imágenes en ZIP: {len(images)}')
        log_lines.append('')
        for product_name, group_rows in groups.items():
            try:
                result = self._create_product_group(product_name, group_rows, images)
                if result.get('product'):
                    stats['created_products'] += 1
                    producto_str = f"✅ Producto: {product_name}"
                    variantes = result.get('variants', [])
                    if variantes:
                        producto_str += f" ({len(variantes)} variantes)"
                        stats['created_variants'] += len(variantes)
                    if result.get('attribute_created'):
                        stats['created_attributes'] += 1
                    log_lines.append(producto_str)
                elif result.get('skipped'):
                    stats['skipped_count'] += 1
                    log_lines.append(f"⏩ Omitido: {product_name} — {result.get('reason', '')}")
            except Exception as e:
                stats['error_count'] += 1
                log_lines.append(f"❌ Error en '{product_name}': {e}")
                _logger.exception('Error importando producto %s', product_name)
        stats['log'] = '\n'.join(log_lines)
        resumen = (
            f"\n{'='*40}\n"
            f"📊 **RESUMEN**\n"
            f"✅ Productos creados: {stats['created_products']}\n"
            f"🔀 Variantes creadas: {stats['created_variants']}\n"
            f"🏷️  Atributos creados: {stats['created_attributes']}\n"
            f"⏩ Omitidos: {stats['skipped_count']}\n"
            f"❌ Errores: {stats['error_count']}\n"
            f"{'='*40}"
        )
        stats['log'] = stats['log'] + '\n' + resumen
        return stats

    def _parse_xlsx(self, xlsx_content):
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_content), data_only=True)
        ws = wb.active
        header = [str(cell.value or '').strip().lower() for cell in ws[1]]
        col_map = self._map_columns(header)
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            vals = {
                'producto': str(row[col_map['producto']] or '').strip() if col_map.get('producto') is not None and row[col_map['producto']] is not None else '',
                'tipo': str(row[col_map['tipo']] or '').strip().lower() if col_map.get('tipo') is not None and row[col_map['tipo']] is not None else '',
                'referencia': str(row[col_map['referencia']] or '').strip() if col_map.get('referencia') is not None and row[col_map['referencia']] is not None else '',
                'precio': row[col_map['precio']] if col_map.get('precio') is not None and row[col_map['precio']] is not None else 0.0,
                'atributo': str(row[col_map['atributo']] or '').strip() if col_map.get('atributo') is not None and row[col_map['atributo']] is not None else '',
                'valor_atributo': str(row[col_map['valor_atributo']] or '').strip() if col_map.get('valor_atributo') is not None and row[col_map['valor_atributo']] is not None else '',
                'imagen': str(row[col_map['imagen']] or '').strip() if col_map.get('imagen') is not None and row[col_map['imagen']] is not None else '',
            }
            if vals['producto']:
                rows.append(vals)
        groups = defaultdict(list)
        for r in rows:
            groups[r['producto']].append(r)
        return dict(groups)

    def _map_columns(self, header):
        col_map = {}
        aliases = {
            'producto': ['producto', 'name', 'nombre', 'product name'],
            'tipo': ['tipo', 'type'],
            'referencia': ['referencia', 'reference', 'default_code', 'código', 'codigo'],
            'precio': ['precio', 'price', 'list_price', 'precio ($)', 'precio (usd)', 'precio($)'],
            'atributo': ['atributo', 'attribute', 'attribute name'],
            'valor_atributo': ['valor atributo', 'valor_atributo', 'attribute value', 'value', 'valor'],
            'imagen': ['imagen', 'image', 'imagen (nombre archivo)', 'imagen(nombre archivo)'],
        }
        for i, h in enumerate(header):
            for key, names in aliases.items():
                if h in names:
                    col_map[key] = i
        return col_map

    def _create_product_group(self, product_name, group_rows, images):
        is_simple = all(r['tipo'] in ('simple', '') for r in group_rows)
        if is_simple:
            return self._create_simple_product(product_name, group_rows[0], images)
        return self._create_variant_product(product_name, group_rows, images)

    def _create_simple_product(self, product_name, row, images):
        ref = row['referencia']
        if ref:
            existing = self.env['product.template'].search([('default_code', '=', ref)], limit=1)
            if existing:
                return {'skipped': True, 'reason': f'default_code "{ref}" ya existe (ID {existing.id})'}
            uom = self.env['uom.uom'].search([('name', '=', 'Unit(s)')], limit=1) or self.env['uom.uom'].search([], limit=1)
            vals = {
                'name': product_name,
                'type': 'consu',
                'default_code': ref or False,
                'list_price': float(row['precio'] or 0.0),
                'uom_id': uom.id if uom else 1,
                'service_tracking': 'no',
                'tracking': 'none',
            }
            if self.categ_id:
                vals['categ_id'] = self.categ_id.id
            product = self.env['product.template'].create(vals)
        if row['imagen'] and row['imagen'] in images:
            product.write({'image_1920': images[row['imagen']]})
        return {'product': product, 'variants': [], 'attribute_created': False}

    def _create_variant_product(self, product_name, group_rows, images):
        padre = next((r for r in group_rows if r['tipo'] == 'padre'), None)
        variantes = [r for r in group_rows if r['tipo'] == 'variante']
        attr_name = padre['atributo'] if padre and padre['atributo'] else (variantes[0]['atributo'] if variantes else '')
        attr_name = attr_name.strip()
        if not attr_name or not variantes:
            return self._create_simple_product(product_name, group_rows[0], images)
        ref_padre = padre['referencia'] if padre else ''
        existing_product = False
        if ref_padre:
            existing_product = self.env['product.template'].search(
                [('default_code', '=', ref_padre)], limit=1
            )
        elif variantes:
            existing_product = self.env['product.template'].search(
                [('name', '=', product_name)], limit=1
            )
        if existing_product and existing_product.attribute_line_ids:
            product_tmpl = existing_product
        else:
            attribute = self._get_or_create_attribute(attr_name)
            unique_values = []
            for r in variantes:
                v = r['valor_atributo'].strip()
                if v and v not in unique_values:
                    unique_values.append(v)
            value_records = []
            for val_name in unique_values:
                value = self._get_or_create_attribute_value(attribute, val_name)
                if value:
                    value_records.append(value.id)
            if not value_records:
                return {'skipped': True, 'reason': 'Sin valores de atributo válidos'}
            uom = self.env['uom.uom'].search([('name', '=', 'Unit(s)')], limit=1) or self.env['uom.uom'].search([], limit=1)
            vals = {
                'name': product_name,
                'type': 'consu',
                'default_code': ref_padre or False,
                'list_price': float(padre['precio'] or 0.0) if padre else 0.0,
                'uom_id': uom.id if uom else 1,
                'service_tracking': 'no',
                'tracking': 'none',
                'attribute_line_ids': [(0, 0, {
                    'attribute_id': attribute.id,
                    'value_ids': [(6, 0, value_records)],
                })],
            }
            if self.categ_id:
                vals['categ_id'] = self.categ_id.id
            product_tmpl = self.env['product.template'].create(vals)
            if padre and padre['imagen'] and padre['imagen'] in images:
                product_tmpl.write({'image_1920': images[padre['imagen']]})
        product_tmpl.invalidate_model()
        product_tmpl._create_variant_ids()
        created_variants = []
        for vr in variantes:
            val_name = vr['valor_atributo'].strip()
            val_ref = vr['referencia']
            try:
                existing_variant = product_tmpl.product_variant_ids.filtered(
                    lambda v: val_name in v.product_template_variant_value_ids.mapped(
                        'product_attribute_value_id.name'
                    )
                )[:1]
                if existing_variant:
                    variant = existing_variant
                    variant.write({
                        'default_code': val_ref or False,
                        'list_price': float(vr['precio'] or 0.0),
                    })
                else:
                    variant = self.env['product.product'].create({
                        'product_tmpl_id': product_tmpl.id,
                        'default_code': val_ref or False,
                        'list_price': float(vr['precio'] or 0.0),
                    })
                if vr['imagen'] and vr['imagen'] in images:
                    variant.write({'image_1920': images[vr['imagen']]})
                created_variants.append(variant)
            except Exception as e:
                _logger.exception('Error creando variante %s - %s', product_name, val_name)
                raise
        return {
            'product': product_tmpl,
            'variants': created_variants,
            'attribute_created': True,
        }

    def _get_or_create_attribute(self, name):
        attribute = self.env['product.attribute'].search([('name', '=', name)], limit=1)
        if not attribute:
            attribute = self.env['product.attribute'].create({
                'name': name,
                'create_variant': 'always',
            })
        return attribute

    def _get_or_create_attribute_value(self, attribute, name):
        value = self.env['product.attribute.value'].search([
            ('name', '=', name),
            ('attribute_id', '=', attribute.id),
        ], limit=1)
        if not value:
            value = self.env['product.attribute.value'].create({
                'name': name,
                'attribute_id': attribute.id,
            })
        return value

    def _extract_zip_images(self, zip_content):
        images = {}
        try:
            with zipfile.ZipFile(io.BytesIO(zip_content)) as zf:
                for name in zf.namelist():
                    base = os.path.basename(name)
                    if not base or base.startswith('.'):
                        continue
                    ext = os.path.splitext(base)[1].lower()
                    if ext not in ('.jpg', '.jpeg', '.png', '.webp', '.gif', '.bmp'):
                        continue
                    data = zf.read(name)
                    images[base] = base64.b64encode(data).decode('utf-8')
        except zipfile.BadZipFile as e:
            _logger.warning('ZIP inválido o vacío: %s', e)
        return images

    def _log_skip(self, msg):
        _logger.info('Skip: %s', msg)
